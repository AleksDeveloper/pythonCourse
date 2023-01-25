from pycountry import countries
import random
import functions

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

wb = Workbook()
#ws = wb.create_sheet(title="Sheet1")
ws = wb.active

global moneyLeft
moneyLeft = 1000000
listNames = []
listMoney = []
listCountries = []

def getRandomMoney(moneyLeftarg):
    global moneyLeft
    randNumber = random.randint(0,(moneyLeftarg - len(listNames)))
    moneyLeft -= randNumber
    return randNumber

def getRandomCountry():
    #randomNumber = random.randint(0,len(countries))
    #print(randomNumber)
    #print(len(countries))
    return(list(countries)[random.randint(0,len(countries))].name)


def isMoneyLeftEnough(money, namesIndexLeft):
    return false

def openList():
    with open('Python24EneroLista.txt') as lista:
        lines = lista.readlines()
    return lines

def fillListMoney():
    for x in range(len(listNames)):
        #print("x in fill list is ",x)
        #print(len(listNames))
        if x == len(listNames)-1:
            listMoney.append(moneyLeft)
        else:
            listMoney.append(getRandomMoney(moneyLeft))

def fillListCountries():
    for x in range(len(listNames)):
        listCountries.append(getRandomCountry())


def setHeaders():
    ws['A1'].value = "Name"
    ws['B1'].value = "Amount"
    ws['C1'].value = "Country"

def setColumnsWidth():
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 25

def setFont():
    font = Font(name='Consolas',size=14,bold=True)
    ws['A1'].font = font
    ws['B1'].font = font
    ws['C1'].font = font

def setAlignment():
    alignment = Alignment(horizontal='center')
    ws['A1'].alignment = alignment
    ws['B1'].alignment = alignment
    ws['C1'].alignment = alignment

def setNameFields():
    index = 2
    for x in listNames:
        ws['A'+str(index)].value = x
        index += 1

def setMoneyFields():
    index = 2
    for x in listMoney:
        ws['B'+str(index)].value = x
        index += 1
    ws['B19'].value = "TOTAL:"
    ws['B19'].alignment = Alignment(horizontal = 'right')
    ws['B20'].value = "=SUM(B2:B18)"

def setCountryFields():
    index = 2
    for x in listCountries:
        ws['C'+str(index)].value = x
        index += 1

def saveDoc():
    wb.save("EXAMPLE.xlsx")


listNames = openList()
#print(listNames)
#print(getRandomMoney(moneyLeft))
#print(getRandomCountry())
setHeaders()
setColumnsWidth()
setFont()
setAlignment()
setNameFields()
fillListMoney()
fillListCountries()
setMoneyFields()
setCountryFields()
functions.holamundo()
saveDoc()

