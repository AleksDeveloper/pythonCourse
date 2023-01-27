import json
import os
import time
from whoisapi import *
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

YOUR_API_KEY = ""
api_key = YOUR_API_KEY
listIPs = []
wb = Workbook()
ws = wb.active

def apiRequest(ip_address, api_keyarg):
    try:
        client = Client(api_key = api_keyarg)
        client.parameters.output_format = 'json'
        info = client.raw_data(ip_address)
        json_object = json.loads(info)
        print("\n",json_object["WhoisRecord"]["domainName"])
        createFile("Jsons/"+str(json_object["WhoisRecord"]["domainName"])+".json", info)
    except Exception as e:
        print("*****An exception ocurred during IP: "+ip_address+" request*****\n"+str(e)+"\n")
        createFileError("badIPs.txt", str(json_object["WhoisRecord"]["domainName"]))

    updateExcelFile(json_object)

def updateExcelFile(json_object):
    path_filename = ""
    domainName = ""
    registrarName = ""
    contactEmail = ""
    registryData_createdDate = ""
    registrant_country = ""

    path_filename = validateRegistry(1, json_object, "Jsons/"+json_object["WhoisRecord"]["domainName"])
    domainName = validateRegistry(2, json_object, path_filename)
    registrarName = validateRegistry(3, json_object, path_filename)
    contactEmail = validateRegistry(4, json_object, path_filename)
    registryData_createdDate = validateRegistry(5, json_object, path_filename)
    registrant_country = validateRegistry(6, json_object, path_filename)


    ws.append([path_filename, domainName, registrarName, contactEmail, registryData_createdDate, registrant_country])
    wb.save("IPSOUTPUT.xlsx")

def validateRegistry(opt, json_object, path):
    if opt == 1:
        try:
            return "Jsons/"+str(json_object["WhoisRecord"]["domainName"])
        except Exception as e:
            print("***ERROR EN REGISTRO "+path+" "+str(e)+" ***")
    elif opt == 2:
        try:
            return str(json_object["WhoisRecord"]["domainName"])
        except Exception as e:
            print("***ERROR EN REGISTRO "+path+" "+str(e)+" ***")
    elif opt == 3:
        try:
            return str(json_object["WhoisRecord"]["registrarName"])
        except Exception as e:
            print("***ERROR EN REGISTRO "+path+" "+str(e)+" ***")
    elif opt == 4:
        try:
            return str(json_object["WhoisRecord"]["contactEmail"])
        except Exception as e:
            print("***ERROR EN REGISTRO "+path+" "+str(e)+" ***")
    elif opt == 5:
        try:
            return str(json_object["WhoisRecord"]["registryData"]["createdDate"])
        except Exception as e:
            print("***ERROR EN REGISTRO "+path+" "+str(e)+" ***")
    elif opt == 6:
        try:
            return str(json_object["WhoisRecord"]["registryData"]["registrant"]["country"])
        except Exception as e:
            print("***ERROR EN REGISTRO "+path+" "+str(e)+" ***")  
    else:
        return ""

def openList(file):
    with open(file) as lista:
        lines = lista.readlines()
    return lines

def fillListIPs(file):
    global listIPs
    listIPs = openList(file)

def createFile(name, value):
    f = open(name, "w")  
    json.dump(value, f, indent = 6)  
    f.close()  

def createFileError(name, value):
    if os.path.exists("badIPs.txt"):
        f = open(name, "a")
        f.write("\n"+str(value))
        f.close()
    else:
        f = open(name, "w")
        f.write(value)
        f.close()

def setHeaders():
    ws['A1'].value = "path filename"
    ws['B1'].value = "domainName"
    ws['C1'].value = "registrarName"
    ws['D1'].value = "contactEmail"
    ws['E1'].value = "registryData.createdDate"
    ws['F1'].value = "registrant.country"

def setColumnsWidth():
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20

setHeaders()
setColumnsWidth()
fillListIPs("ips.txt")
print(len(listIPs))
for x in range(50):
    apiRequest(str(listIPs[x].strip()), api_key)
    print("Request de: "+listIPs[x])
    time.sleep(1)
#apiRequest("141.202.225.1",api_key)
#apiRequest("126.151.90.178",api_key)

#print(listIPs)
#apiRequest(listIPs[3], api_key)
#apiRequest("29.79.215.228", api_key)
#apiRequest(str(listIPs[0].strip()), api_key)